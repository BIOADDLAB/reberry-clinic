#!/usr/bin/env python3
"""Normalize the clinic XLSX price table into Firestore seed JSON."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.cell.cell import MergedCell


SHEETS = [
    ("보톡스,제모", "botox-hair-removal", "보톡스·제모", [(27, 27)]),
    ("필러", "filler", "필러·실리프팅", [(18, 35)]),
    ("티타늄", "titanium", "티타늄", []),
    ("co2,켈로이드,지분 ", "co2-keloid-contouring", "CO2·켈로이드·지방분해", []),
    ("스킨부스터", "skin-booster", "스킨부스터", []),
    ("문신,바디토닝", "tattoo-body-toning", "문신제거·바디토닝", []),
    ("포텐자,여드름,모공흉터", "potenza-acne-scar", "포텐자·여드름·모공흉터", [(11, 26)]),
    ("리프팅", "lifting", "리프팅", []),
    ("관리", "skin-care", "피부관리", []),
    ("수액", "iv-therapy", "수액·주사", []),
]

TARGET_CATEGORIES = [
    ("lifting-laser", "리프팅 레이저"),
    ("filler-face-volume", "필러·페이스볼륨"),
    ("anti-aging", "안티에이징"),
    ("skin-booster", "스킨부스터"),
    ("pigment-pore-acne", "색소·모공·여드름"),
    ("toxin-contour", "톡신·윤곽"),
    ("skin-care", "스킨케어"),
    ("body-line", "바디라인"),
    ("hair-removal", "제모"),
]

# 원본 시트의 세부 유형은 이름 정제에 계속 사용하고, 화면에서는 위 대분류로 묶는다.
KIND_LABELS = {
    "botox": "보톡스",
    "hair-removal": "제모",
    "filler": "필러",
    "thread-lifting": "실리프팅",
    "lifting": "리프팅",
    "skin-booster": "스킨부스터",
    "acne-ptt": "여드름",
    "co2": "CO2",
    "keloid": "켈로이드",
    "pore-scar": "모공흉터",
    "potenza": "포텐자",
    "pigment": "색소",
    "skin-care": "피부관리",
    "tattoo-body-toning": "문신제거·바디토닝",
    "body-contouring": "지방분해",
    "iv-therapy": "수액",
    "redness": "홍조",
}

SECTION_DEFS = [
    ("lifting-laser", "titanium", "티타늄"),
    ("lifting-laser", "onda", "온다"),
    ("lifting-laser", "v-ro", "브이로"),
    ("lifting-laser", "levinas", "레비나스"),
    ("lifting-laser", "inmode", "인모드"),
    ("lifting-laser", "shurink", "슈링크"),
    ("filler-face-volume", "volume", "필러(볼륨)"),
    ("filler-face-volume", "special", "특수부위 필러"),
    ("filler-face-volume", "package", "필러 패키지"),
    ("filler-face-volume", "nose-line", "아름다운 코라인"),
    ("filler-face-volume", "hyalase", "히알라제"),
    ("anti-aging", "dews", "듀스 실리프팅"),
    ("anti-aging", "iv-therapy", "수액·주사"),
    ("skin-booster", "skin-booster", "스킨부스터"),
    ("pigment-pore-acne", "pigment-package", "색소·미백 패키지"),
    ("pigment-pore-acne", "redness-package", "홍조 패키지"),
    ("pigment-pore-acne", "acne-package", "여드름 패키지"),
    ("pigment-pore-acne", "gold-ptt", "골드 PTT"),
    ("pigment-pore-acne", "potenza", "포텐자"),
    ("pigment-pore-acne", "pore-package", "모공·흉터 패키지"),
    ("pigment-pore-acne", "co2", "CO2"),
    ("pigment-pore-acne", "keloid", "켈로이드"),
    ("pigment-pore-acne", "tattoo-removal", "문신제거"),
    ("toxin-contour", "botox", "보톡스"),
    ("skin-care", "skin-care", "피부관리"),
    ("body-line", "body-contouring", "바디 지방분해 주사"),
    ("body-line", "body-toning", "바디 토닝"),
    ("hair-removal", "female", "여성제모"),
    ("hair-removal", "male", "남성제모"),
]
SECTION_LABELS = {f"{category_id}--{slug}": label for category_id, slug, label in SECTION_DEFS}
COUNT_LABEL_RE = re.compile(r"^\d+(?:\.\d+)?\s*(?:회|개|병)$")

PRICE_FIXES = {
    "699,00원": "699,000원",
    "14,9000원": "149,000원",
    "279,00원": "279,000원",
    "199000원": "199,000원",
}
PACKAGE_RE = re.compile(r"pkg|패키지", re.IGNORECASE)
PRICE_RE = re.compile(r"(?:(\(\d+\s*회\))\s*)?(\d[\d,\s]*)\s*원\s*$")
OPTION_ONLY_RE = re.compile(
    r"^(?:\d+(?:\.\d+)?\s*(?:회|cc|u|샷|줄|kj)|\d+\s*/\s*\d+\s*회|"
    r"1회|3회|4회|5회|6회|8회|10회|12회|본원|타원|국산|수입|x)$",
    re.IGNORECASE,
)


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(value).replace("\r", "\n")).strip()


def parse_price(value: Any) -> tuple[int, str] | None:
    if isinstance(value, (int, float)) and value >= 1_000:
        return int(value), ""
    text = clean(value)
    fixed = PRICE_FIXES.get(text.replace(" ", ""), text)
    match = PRICE_RE.fullmatch(fixed)
    if not match:
        return None
    amount = int(re.sub(r"\D", "", match.group(2)))
    prefix = clean(match.group(1))
    return amount, prefix


def stable_id(*parts: str) -> str:
    source = "\0".join(parts)
    return hashlib.sha1(source.encode("utf-8")).hexdigest()[:20]


def merged_context(ws: Any) -> dict[tuple[int, int], tuple[str, str]]:
    result: dict[tuple[int, int], tuple[str, str]] = {}
    for merged_range in ws.merged_cells.ranges:
        anchor = ws.cell(merged_range.min_row, merged_range.min_col)
        value = clean(anchor.value)
        if not value:
            continue
        coordinate = anchor.coordinate
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                result[(row, col)] = (value, coordinate)
    return result


def context_value(ws: Any, merged: dict[tuple[int, int], tuple[str, str]], row: int, col: int) -> tuple[str, str]:
    if (row, col) in merged:
        return merged[(row, col)]
    cell = ws.cell(row, col)
    return clean(cell.value), cell.coordinate


def excluded(row: int, ranges: list[tuple[int, int]]) -> bool:
    return any(start <= row <= end for start, end in ranges)


def source_row(item: dict) -> int:
    match = re.search(r"\d+", item["sourceCell"])
    return int(match.group()) if match else 0


def target_category(item: dict) -> str:
    if "targetCategoryId" in item:
        return item["targetCategoryId"]
    source = item["categoryId"]
    row = source_row(item)
    if source == "botox-hair-removal":
        return "botox" if row <= 26 else "hair-removal"
    if source == "filler":
        return "thread-lifting" if row >= 42 else "filler"
    if source in {"titanium", "lifting"}:
        return "lifting"
    if source == "co2-keloid-contouring":
        if row <= 12:
            return "co2"
        if row <= 19:
            return "keloid"
        return "body-contouring"
    if source == "skin-booster":
        return "skin-booster"
    if source == "tattoo-body-toning":
        return "tattoo-body-toning"
    if source == "potenza-acne-scar":
        return "acne-ptt" if row >= 29 else "potenza"
    if source == "skin-care":
        return "skin-care"
    if source == "iv-therapy":
        return "iv-therapy"
    raise RuntimeError(f"unmapped source category: {source}")


def package_item(
    category_id: str,
    sheet: str,
    cell: str,
    name: str,
    options: list[tuple[str, int, str]],
    description: str = "",
) -> dict:
    item_id = stable_id(category_id, sheet, cell, name)
    return {
        "docId": item_id,
        "categoryId": category_id,
        "targetCategoryId": category_id,
        "section": KIND_LABELS[category_id],
        "name": name,
        "description": description,
        "options": [
            {
                "id": stable_id(item_id, source_cell, label),
                "label": label,
                "price": price,
                "sourceCell": source_cell,
            }
            for label, price, source_cell in options
        ],
        "sort": 0,
        "isPublished": True,
        "sourceSheet": sheet,
        "sourceCell": cell,
    }


def package_items() -> list[dict]:
    return [
        package_item("botox", "보톡스,제모", "A27", "원데이 종아리 알빼기 PKG", [("기본", 179_000, "D27")]),
        package_item(
            "filler",
            "필러",
            "A19",
            "풀페이스 필러 10cc",
            [
                ("국산 · 레나/아띠", 890_000, "C20"),
                ("국산 프리미엄 · 르네필/리쥬비엘/레나", 1_500_000, "D20"),
                ("수입 · 레스틸렌/벨로테로", 2_500_000, "E20"),
            ],
        ),
        package_item(
            "filler",
            "필러",
            "A21",
            "풀페이스 필러 20cc",
            [
                ("국산 · 레나/아띠", 1_650_000, "C21"),
                ("국산 프리미엄 · 르네필/리쥬비엘/레나", 2_500_000, "D21"),
                ("수입 · 레스틸렌/벨로테로", 4_400_000, "E21"),
            ],
        ),
        package_item(
            "filler",
            "필러",
            "A23",
            "원데이 목주름 PKG",
            [("기본", 590_000, "E24")],
            "브이로 200샷 + 벨로테로 주름필러 1CC + 스킨보톡스 4CC",
        ),
        package_item(
            "filler",
            "필러",
            "A26",
            "원데이 팔자주름 PKG",
            [("기본", 390_000, "E27")],
            "브이로 200샷 + 아띠에르 팔자필러 2CC + 레나 주름필러 1CC",
        ),
        package_item(
            "filler",
            "필러",
            "A29",
            "원데이 팔자주름 스페셜 PKG",
            [("기본", 690_000, "E30")],
            "온다 50KJ + 브이로 200샷 + 아띠에르 팔자필러 2CC + 레나 주름필러 1CC",
        ),
        package_item(
            "filler",
            "필러",
            "A32",
            "아름다운 코라인",
            [
                ("아띠에르", 539_000, "E33"),
                ("리쥬비엘 C", 649_000, "E34"),
                ("레스틸렌 리프트", 849_000, "E35"),
            ],
            "코 필러 1CC + 하이코 2줄 + 바비코 2줄",
        ),
        package_item(
            "acne-ptt",
            "포텐자,여드름,모공흉터",
            "A12",
            "안티아크네 8주",
            [("기본", 1_190_000, "J12")],
            "매 회차 압출·염증주사 + 격주 골드PTT + 제네시스",
        ),
        package_item(
            "acne-ptt",
            "포텐자,여드름,모공흉터",
            "A14",
            "프리미엄 안티아크네 8주",
            [("기본", 1_490_000, "J14")],
            "격주 골드PTT + LDM 4회 + 포텐자 아그네스 + 청색LED 4회",
        ),
        package_item(
            "acne-ptt",
            "포텐자,여드름,모공흉터",
            "A16",
            "시그니처 안티아크네 12주",
            [("기본", 2_590_000, "J16")],
            "포텐자 N25팁&엑소좀 3회(4주 간격) + LDM&엑소좀 5회(2주 간격)",
        ),
        package_item(
            "pore-scar",
            "포텐자,여드름,모공흉터",
            "A19",
            "도자기 피부 패키지",
            [("4회", 1_590_000, "J20"), ("8회", 2_490_000, "J21")],
            "피코프락셀 + 포텐자N팁&쥬베룩 + 1·3회차(8회는 1·3·5·7회차) 나비존 스킨보톡스",
        ),
        package_item(
            "pore-scar",
            "포텐자,여드름,모공흉터",
            "A22",
            "프리미엄 도자기 8회",
            [("기본", 2_990_000, "I23")],
            "피코프락셀 8회 + 포텐자N팁&쥬베룩 8회 + 크라이오 4회 + 시너젯/쥬베룩 3회",
        ),
        package_item(
            "pore-scar",
            "포텐자,여드름,모공흉터",
            "A25",
            "시그니처 도자기 8회",
            [("기본", 3_490_000, "I26")],
            "피코프락셀 8회 + 포텐자N팁&쥬베룩 8회 + 크라이오 8회 + 울킨 8회 + 시너젯/쥬베룩/서브시전 4회",
        ),
        package_item(
            "pigment",
            "색소pkg",
            "A2",
            "진주광채 피코&헐리우드 토닝 4회",
            [("기본", 1_290_000, "I2")],
            "헐리우드 토닝 4회 + 제네시스 콜라겐 레이저 2회 + 리베리 시그니처 스킨부스터 2회 + LDM 2회",
        ),
        package_item(
            "pigment",
            "색소pkg",
            "A4",
            "프리미엄 진주광채 피코&헐리우드 토닝 8회",
            [("기본", 1_990_000, "I4")],
            "헐리우드 토닝 8회 + 제네시스·LDM·이온토관리(미백) 4회 + 리베리 시그니처 스킨부스터 4회 + 맞춤레이저 2회(리팟 포함)",
        ),
        package_item(
            "pigment",
            "색소pkg",
            "A6",
            "시그니처 진주광채 피코&헐리우드 토닝 8회",
            [("기본", 2_490_000, "I6")],
            "헐리우드 토닝 8회 + 제네시스 4회 + 맞춤레이저 2회(리팟 포함) + 격주 리베리 시그니처 스킨부스터 4회 + 포텐자 N팁&시그니처 스킨부스터 2회 + 점/검버섯 제거 20개 + 울킨 4회 + 이온토관리 4회",
        ),
        package_item(
            "pigment",
            "색소pkg",
            "A8",
            "프레스티지 진주광채 피코&헐리우드 토닝 10회",
            [("기본", 2_790_000, "I8")],
            "헐리우드 토닝 10회 + 제네시스 4회 + 맞춤레이저 2회(리팟 포함) + 격주 리베리 시그니처 스킨부스터 4회 + 포텐자 N팁&시그니처 스킨부스터 4회 + 점/검버섯 제거 20개 + 울킨 4회 + 맞춤 피부관리 4회",
        ),
        package_item(
            "pigment",
            "색소pkg",
            "A10",
            "플래티넘 진주광채 피코&헐리우드 토닝 12회",
            [("기본", 3_190_000, "I10")],
            "헐리우드 토닝 12회 + 제네시스 4회 + 맞춤레이저 4회(리팟 포함) + 리베리 시그니처 스킨부스터 4회 + 포텐자 N팁&시그니처 스킨부스터 4회 + 점/검버섯 제거 20개 + 울킨 4회 + 맞춤 피부관리 4회",
        ),
        package_item(
            "redness",
            "홍조",
            "A2",
            "엑셀 홍조",
            [("1회", 299_000, "G3"), ("4회", 990_000, "G4")],
            "매 회차 제네시스 + 엑셀브이 + 진정관리",
        ),
        package_item(
            "redness",
            "홍조",
            "A5",
            "프리미엄 엑셀 홍조",
            [("4회", 1_690_000, "G7"), ("8회", 2_690_000, "G8")],
            "제네시스 + 엑셀브이 + 진정관리 + 포텐자N25팁 + 쥬베룩스킨",
        ),
        package_item(
            "redness",
            "홍조",
            "A9",
            "프리미엄 재생 홍조",
            [("1회", 319_000, "G10"), ("4회", 1_490_000, "G11"), ("8회", 2_290_000, "G12")],
            "1회: 제네시스 + 포텐자N25 + 물광 + 스킨보톡스\n4·8회: 골드PTT + 울킨 + 제네시스 + 스킨보톡스 / 포텐자N25 + 물광",
        ),
    ]


def section_id_for(item: dict, category_id: str) -> str:
    sheet = item["sourceSheet"].strip()
    row = source_row(item)
    col = re.match(r"[A-Z]+", item["sourceCell"]).group()
    if category_id == "botox":
        slug = "botox"
    elif category_id == "hair-removal":
        slug = "female" if row < 47 else "male"
    elif category_id == "filler":
        if row <= 8:
            slug = "volume"
        elif row <= 17:
            slug = "special"
        elif row <= 31:
            slug = "package"
        elif row <= 38:
            slug = "nose-line"
        else:
            slug = "hyalase"
    elif category_id == "thread-lifting":
        slug = "dews"
    elif category_id == "lifting":
        if sheet == "티타늄":
            slug = "titanium"
        elif row <= 10:
            slug = "onda" if col < "F" else "v-ro"
        elif row <= 15:
            slug = "levinas" if col < "F" else "v-ro"
        elif row <= 27:
            slug = "inmode"
        else:
            slug = "shurink"
    elif category_id == "skin-booster":
        slug = "skin-booster"
    elif category_id == "acne-ptt":
        slug = "gold-ptt" if row >= 29 else "acne-package"
    elif category_id == "co2":
        slug = "co2"
    elif category_id == "keloid":
        slug = "keloid"
    elif category_id == "pore-scar":
        slug = "pore-package"
    elif category_id == "potenza":
        slug = "potenza"
    elif category_id == "pigment":
        slug = "pigment-package"
    elif category_id == "skin-care":
        slug = "skin-care"
    elif category_id == "tattoo-body-toning":
        slug = "body-toning" if row <= 4 else "tattoo-removal"
    elif category_id == "body-contouring":
        slug = "body-contouring"
    elif category_id == "iv-therapy":
        slug = "iv-therapy"
    elif category_id == "redness":
        slug = "redness-package"
    else:
        raise RuntimeError(f"missing section mapping: {category_id}")

    if category_id == "lifting":
        group_id = "lifting-laser"
    elif category_id == "filler":
        group_id = "filler-face-volume"
    elif category_id in {"thread-lifting", "iv-therapy"}:
        group_id = "anti-aging"
    elif category_id == "skin-booster":
        group_id = "skin-booster"
    elif category_id in {"pigment", "redness", "acne-ptt", "potenza", "pore-scar", "co2", "keloid"}:
        group_id = "pigment-pore-acne"
    elif category_id == "tattoo-body-toning":
        group_id = "body-line" if slug == "body-toning" else "pigment-pore-acne"
    elif category_id == "botox":
        group_id = "toxin-contour"
    elif category_id == "skin-care":
        group_id = "skin-care"
    elif category_id == "body-contouring":
        group_id = "body-line"
    elif category_id == "hair-removal":
        group_id = "hair-removal"
    else:
        raise RuntimeError(f"missing category group mapping: {category_id}")
    return f"{group_id}--{slug}"


def complete_name(item: dict, category_id: str, section_id: str) -> str:
    name = clean(item["name"])
    row = source_row(item)
    if category_id == "botox":
        parents = {
            13: "사각턱 보톡스", 14: "사각턱 보톡스",
            17: "침샘 보톡스", 18: "침샘 보톡스",
            22: "바디 보톡스", 23: "바디 보톡스",
            26: "스킨보톡스",
        }
        if row in parents:
            name = f"{parents[row]} {name}"
        name = re.sub(r"^주름(?=\s|3|올인원)", "주름보톡스 ", name).replace("  ", " ")
        name = name.replace("리터치 사각턱", "사각턱 보톡스 리터치")
        name = name.replace("침샘 25U 추가", "침샘 보톡스 25U 추가")
        if "보톡스" not in name and "PKG" not in name:
            name = f"{name} 보톡스"
    elif category_id == "filler" and section_id.endswith("--special"):
        if "특수부위 필러" not in name:
            name = f"특수부위 필러 {name}"
    elif category_id == "thread-lifting":
        if name == "듀스 실리프팅":
            name = "듀스 실리프팅"
        elif not name.startswith("듀스"):
            name = f"듀스 실리프팅 {name}"
    elif category_id == "lifting":
        section = SECTION_LABELS[section_id]
        if not name.lower().startswith(section.lower()) and section not in name:
            name = f"{section} {name}"
        if section in {"온다", "브이로", "레비나스", "인모드", "슈링크"} and "리프팅" not in name:
            name = f"{name} 리프팅"
    elif category_id == "co2":
        col = re.match(r"[A-Z]+", item["sourceCell"]).group()
        if col == "D" and row in {7, 8, 9}:
            name = "점 제거 (소 사이즈 기준)"
        elif row == 12:
            name = "특수 사마귀 (1cm~)"
        elif name.startswith("CO2 "):
            name = name[4:]
    elif category_id == "keloid":
        name = re.sub(r"^켈로이드(?: 주사)?\s*", "", name)
    elif category_id == "acne-ptt" and "골드" in name and "PTT" in name.upper():
        name = "골드 PTT"
    elif category_id == "potenza" and not name.startswith("포텐자"):
        name = f"포텐자 {name}"
    elif category_id == "tattoo-body-toning" and section_id.endswith("--body-toning") and row == 4:
        name = "바디토닝 패키지"
    return clean(name)


def session_label_for(item: dict, option: dict, section_id: str) -> str:
    label = clean(option["label"])
    if COUNT_LABEL_RE.fullmatch(label):
        return label.replace(" ", "")
    if label.lower() == "기본":
        if section_id.endswith("--co2"):
            count_match = re.match(r"(\d+개)", item["name"])
            return count_match.group(1) if count_match else "1개"
        trailing_count = re.search(r"(\d+회)\s*$", item["name"])
        return trailing_count.group(1) if trailing_count else "1회"
    return "1회"


def split_patient_items(raw_item: dict, category_id: str, section_id: str) -> list[dict]:
    base_name = complete_name(raw_item, category_id, section_id)
    options = raw_item["options"]
    split_each = category_id in {"botox", "filler"} or any(
        clean(option["label"]).lower() != "기본" and not COUNT_LABEL_RE.fullmatch(clean(option["label"]))
        for option in options
    )

    groups: list[tuple[list[dict], str]] = []
    if split_each:
        for option in options:
            label = clean(option["label"])
            variants = [label]
            if category_id == "filler" and "/" in label:
                prefix, product_part = ("", label)
                if "·" in label:
                    prefix, product_part = [clean(part) for part in label.rsplit("·", 1)]
                products = [clean(product) for product in product_part.split("/") if clean(product)]
                variants = [f"{prefix} · {product}".strip(" ·") for product in products]
            groups.extend([([option], variant) for variant in variants])
    else:
        groups = [(options, "")]
    result = []
    for group_index, (group, variant_override) in enumerate(groups):
        name = base_name
        product_label = ""
        if split_each:
            variant = variant_override or clean(group[0]["label"])
            if category_id == "filler":
                parts = [clean(part) for part in variant.split("·")]
                volume = next((part for part in parts if re.fullmatch(r"\d+(?:\.\d+)?cc", part, re.I)), "")
                product = " · ".join(part for part in parts if part != volume)
                if product and product.lower() != "기본":
                    product_label = product
                if section_id.endswith("--volume"):
                    name = f"볼륨 필러 {base_name}"
                if volume and volume.lower() not in name.lower():
                    name = f"{name} {volume}"
                if section_id.endswith("--volume") and "필러" not in name:
                    name = f"{name} 볼륨 필러"
            elif category_id == "botox":
                if variant.lower() != "기본":
                    product_label = variant
            elif variant.lower() != "기본":
                name = f"{base_name} {variant}"

        sessions = [
            {
                "id": stable_id(category_id, section_id, raw_item["sourceCell"], option["sourceCell"], "session"),
                "label": session_label_for(raw_item, option, section_id),
                "price": option["price"],
                "sourceCell": option["sourceCell"],
            }
            for option in group
        ]
        trailing = sessions[0]["label"] if len(sessions) == 1 else ""
        if trailing and trailing != "1회" and re.search(rf"\s*{re.escape(trailing)}\s*$", name):
            name = re.sub(rf"\s*{re.escape(trailing)}\s*$", "", name).strip()
        if category_id == "thread-lifting" and clean(group[0]["label"]).lower() != "기본":
            name = f"{base_name} {clean(group[0]['label'])}"

        item_id = stable_id(
            category_id,
            section_id,
            raw_item["sourceSheet"],
            raw_item["sourceCell"],
            name,
            product_label,
            str(group_index),
        )
        result.append(
            {
                "docId": item_id,
                "categoryId": category_id,
                "sectionId": section_id,
                "name": clean(name),
                "productLabel": clean(product_label),
                "description": clean(raw_item.get("description", ""))
                or (
                    {
                        5: "이마, 미간, 눈썹 위, 눈가, 자갈턱, 콧등 중 1부위",
                        9: "팔자, 치마주름, 입꼬리, 콧볼, 눈밑 중 1부위",
                    }.get(source_row(raw_item), "")
                    if category_id == "botox"
                    else "약처방전 포함 · 압출, TA 별도"
                    if section_id.endswith("--gold-ptt")
                    else ""
                ),
                "sessions": sessions,
                "sort": 0,
                "isPublished": True,
                "sourceSheet": raw_item["sourceSheet"],
                "sourceCell": raw_item["sourceCell"],
            }
        )
    return result


def normalize_sheet(ws: Any, category_id: str, category_label: str, excluded_ranges: list[tuple[int, int]]) -> list[dict]:
    merged = merged_context(ws)
    headers: dict[int, str] = {}
    grouped: "OrderedDict[str, dict]" = OrderedDict()
    fallback_by_column: dict[int, str] = {}
    last_single_title: tuple[int, str, str] | None = None

    for row_number in range(1, ws.max_row + 1):
        if excluded(row_number, excluded_ranges):
            continue

        price_cells = []
        for cell in ws[row_number]:
            if isinstance(cell, MergedCell):
                continue
            parsed = parse_price(cell.value)
            if parsed:
                price_cells.append((cell, parsed))

        if not price_cells:
            row_texts = []
            for col in range(1, ws.max_column + 1):
                value, _ = context_value(ws, merged, row_number, col)
                if value and not parse_price(value) and value not in row_texts:
                    row_texts.append(value)
            if len(row_texts) >= 2:
                for col in range(1, ws.max_column + 1):
                    value, _ = context_value(ws, merged, row_number, col)
                    if value and not parse_price(value):
                        headers[col] = value
            elif len(row_texts) == 1:
                value, coordinate = context_value(ws, merged, row_number, 1)
                if not value:
                    for col in range(2, ws.max_column + 1):
                        value, coordinate = context_value(ws, merged, row_number, col)
                        if value:
                            break
                last_single_title = (row_number, value, coordinate)
            continue

        for price_cell, (amount, inline_option) in price_cells:
            left_values: list[tuple[int, str, str]] = []
            seen_content = False
            for col in range(price_cell.column - 1, 0, -1):
                value, coordinate = context_value(ws, merged, row_number, col)
                if not value:
                    if seen_content:
                        break
                    continue
                seen_content = True
                if not parse_price(value):
                    left_values.insert(0, (col, value, coordinate))

            deduped_left: list[tuple[int, str, str]] = []
            for entry in left_values:
                if not any(existing[2] == entry[2] for existing in deduped_left):
                    deduped_left.append(entry)
            left_values = deduped_left

            option_parts: list[str] = []
            name_entry = None
            for entry in reversed(left_values):
                if OPTION_ONLY_RE.fullmatch(entry[1].replace(" ", "")) and name_entry is None:
                    option_parts.insert(0, entry[1])
                    continue
                name_entry = entry
                break

            if name_entry is None:
                all_left_values: list[tuple[int, str, str]] = []
                for col in range(1, price_cell.column):
                    value, coordinate = context_value(ws, merged, row_number, col)
                    if value and not parse_price(value) and not any(entry[2] == coordinate for entry in all_left_values):
                        all_left_values.append((col, value, coordinate))
                name_entry = next(
                    (
                        entry
                        for entry in reversed(all_left_values)
                        if not OPTION_ONLY_RE.fullmatch(entry[1].replace(" ", ""))
                    ),
                    None,
                )

            if name_entry is None:
                previous_title_is_name = (
                    last_single_title
                    and last_single_title[0] == row_number - 1
                    and not OPTION_ONLY_RE.fullmatch(last_single_title[1].replace(" ", ""))
                )
                if previous_title_is_name:
                    name_entry = (1, last_single_title[1], last_single_title[2])
                elif left_values:
                    count_only = re.fullmatch(r"\d+\s*회", left_values[-1][1].replace(" ", ""))
                    fallback_key = fallback_by_column.get(price_cell.column)
                    if count_only and fallback_key and fallback_key in grouped:
                        name_entry = (
                            1,
                            grouped[fallback_key]["name"],
                            grouped[fallback_key]["sourceCell"],
                        )
                    else:
                        name_entry = next(
                            (entry for entry in left_values if entry[1].strip().lower() != "x"),
                            left_values[-1],
                        )
                        option_parts = [part for part in option_parts if part != name_entry[1]]
                else:
                    fallback_key = fallback_by_column.get(price_cell.column)
                    if fallback_key and fallback_key in grouped:
                        name_entry = (
                            1,
                            grouped[fallback_key]["name"],
                            grouped[fallback_key]["sourceCell"],
                        )
                    else:
                        name_entry = (1, f"{category_label} {row_number}행", f"A{row_number}")

            name = name_entry[1]
            name_anchor = name_entry[2]
            if PACKAGE_RE.search(name):
                continue

            if inline_option:
                option_parts.append(inline_option)
            header = headers.get(price_cell.column, "")
            if header and header != name and header not in option_parts:
                option_parts.append(header)
            unique_options = []
            for part in option_parts:
                if part and part.lower() != "x" and part not in unique_options:
                    unique_options.append(part)
            option_label = " · ".join(unique_options) or "기본"

            group_key = f"{category_id}:{name_anchor}"
            if group_key not in grouped:
                grouped[group_key] = {
                    "docId": stable_id(category_id, name_anchor, name),
                    "categoryId": category_id,
                    "section": category_label,
                    "name": name,
                    "description": "",
                    "options": [],
                    "sort": len(grouped),
                    "isPublished": True,
                    "sourceSheet": ws.title.strip(),
                    "sourceCell": name_anchor,
                }
            item = grouped[group_key]
            option_id = stable_id(item["docId"], price_cell.coordinate, option_label)
            if not any(option["id"] == option_id for option in item["options"]):
                item["options"].append(
                    {
                        "id": option_id,
                        "label": option_label,
                        "price": amount,
                        "sourceCell": price_cell.coordinate,
                    }
                )
            fallback_by_column[price_cell.column] = group_key

    return [item for item in grouped.values() if item["options"]]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.source, data_only=True)
    now_marker = "xlsx-seed-v4"
    categories = [
        {
            "docId": category_id,
            "label": label,
            "sort": sort,
            "isPublished": True,
            "seedVersion": now_marker,
        }
        for sort, (category_id, label) in enumerate(TARGET_CATEGORIES)
    ]
    raw_items = []

    for sheet_name, category_id, label, excluded_ranges in SHEETS:
        ws = workbook[sheet_name]
        sheet_items = normalize_sheet(ws, category_id, label, excluded_ranges)
        raw_items.extend(sheet_items)

    raw_items.extend(package_items())
    sheet_order = {sheet_name.strip(): index for index, sheet_name in enumerate(workbook.sheetnames)}
    raw_items.sort(
        key=lambda item: (
            sheet_order.get(item["sourceSheet"].strip(), len(sheet_order)),
            source_row(item),
            openpyxl.utils.column_index_from_string(re.match(r"[A-Z]+", item["sourceCell"]).group()),
        )
    )

    items = []
    for raw_item in raw_items:
        category_id = target_category(raw_item)
        section_id = section_id_for(raw_item, category_id)
        for item in split_patient_items(raw_item, category_id, section_id):
            item["categoryId"] = section_id.split("--", 1)[0]
            item["seedVersion"] = now_marker
            items.append(item)

    consolidated: list[dict] = []
    by_name: dict[tuple[str, str, str], dict] = {}
    for item in items:
        key = (item["sectionId"], item["name"], item.get("productLabel", ""))
        existing = by_name.get(key)
        existing_labels = {session["label"] for session in existing["sessions"]} if existing else set()
        incoming_labels = {session["label"] for session in item["sessions"]}
        if existing and existing_labels.isdisjoint(incoming_labels):
            existing["sessions"].extend(item["sessions"])
        else:
            consolidated.append(item)
            by_name[key] = item
    items = consolidated

    section_counts: dict[str, int] = {}
    for item in items:
        item["sort"] = section_counts.get(item["sectionId"], 0)
        section_counts[item["sectionId"]] = item["sort"] + 1

    sections = [
        {
            "docId": f"{category_id}--{slug}",
            "categoryId": category_id,
            "label": label,
            "sort": sort,
            "isPublished": True,
            "seedVersion": now_marker,
        }
        for category_id, category_sections in OrderedDict(
            (
                category_id,
                [(slug, label) for candidate, slug, label in SECTION_DEFS if candidate == category_id],
            )
            for category_id, _ in TARGET_CATEGORIES
        ).items()
        for sort, (slug, label) in enumerate(category_sections)
    ]

    for item in items:
        item["seedVersion"] = now_marker

    invalid_prices = [
        (item["name"], session["price"])
        for item in items
        for session in item["sessions"]
        if not isinstance(session["price"], int) or session["price"] <= 0
    ]
    invalid_session_labels = [
        (item["name"], session["label"])
        for item in items
        for session in item["sessions"]
        if not COUNT_LABEL_RE.fullmatch(session["label"])
    ]
    missing_sections = [item["name"] for item in items if item["sectionId"] not in SECTION_LABELS]
    source_price_mismatches = []
    for item in items:
        if item["sourceSheet"] not in workbook.sheetnames:
            continue
        source_ws = workbook[item["sourceSheet"]]
        for session in item["sessions"]:
            parsed = parse_price(source_ws[session["sourceCell"]].value)
            if not parsed or parsed[0] != session["price"]:
                source_price_mismatches.append(
                    (item["name"], session["sourceCell"], session["price"], parsed[0] if parsed else None)
                )
    doc_ids = [item["docId"] for item in items]
    duplicate_ids = len(doc_ids) - len(set(doc_ids))

    if invalid_prices or invalid_session_labels or missing_sections or source_price_mismatches or duplicate_ids:
        raise RuntimeError(
            "validation failed: "
            f"invalid_prices={invalid_prices}, invalid_session_labels={invalid_session_labels}, "
            f"missing_sections={missing_sections}, source_price_mismatches={source_price_mismatches}, "
            f"duplicate_ids={duplicate_ids}"
        )

    payload = {
        "version": now_marker,
        "source": unicodedata.normalize("NFC", args.source.name),
        "categories": categories,
        "sections": sections,
        "items": items,
        "summary": {
            "categoryCount": len(categories),
            "sectionCount": len(sections),
            "itemCount": len(items),
            "sessionCount": sum(len(item["sessions"]) for item in items),
        },
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload["summary"], ensure_ascii=False))


if __name__ == "__main__":
    main()
